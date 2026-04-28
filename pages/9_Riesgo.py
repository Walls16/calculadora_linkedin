"""
pages/9_Riesgo.py
-----------------
Módulo 9: Análisis de Riesgo.
Tab 1: Riesgo de Portafolios de Acciones — VaR / CVaR paramétrico y Monte Carlo.
Tab 2: CreditMetrics — Riesgo de crédito en portafolios de bonos corporativos.
"""

import datetime
import io

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from scipy.stats import norm

from credit_engine import (
    RATINGS_EMIT as RATINGS,
    RATING_IDX,
    DEFAULT_TM,
    DEFAULT_SPREADS,
    DEFAULT_TREASURY,
    _TM_RAW_17x19,
    build_transition_matrix,
    bond_values_per_rating,
    independent_distribution,
    var_cvar_from_distribution,
    var_cvar_parametric,             # ← VaR Normal (coincide con Excel)
    var_cvar_from_simulations,
    scale_var_cvar,
    gaussian_copula_simulation,
    thresholds_per_bond,
    expected_value_and_sigma,
    N_EMIT as N_R,
    TRADING_DAYS,
)
from utils import (
    get_engine, page_header, paso_a_paso, separador,
    themed_info, themed_success, themed_warning, themed_error,
    plotly_theme, plotly_colors, get_current_theme,
)

# =============================================================================
# CONFIGURACIÓN
# =============================================================================
st.set_page_config(
    page_title="Riesgo · Calculadora Financiera",
    page_icon="📉",
    layout="wide",
)

engine = get_engine()

page_header(
    titulo="9. Análisis de Riesgo",
    subtitulo="Portafolios de Acciones (VaR/CVaR) · CreditMetrics para Bonos Corporativos"
)

# =============================================================================
# TABS PRINCIPALES
# =============================================================================
tab_port, tab_cm = st.tabs([
    "Riesgo de Portafolios (Acciones)",
    "CreditMetrics — Riesgo de Bonos",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — RIESGO DE PORTAFOLIOS (ACCIONES)
# ══════════════════════════════════════════════════════════════════════════════
with tab_port:
    st.markdown("### Análisis de Valor en Riesgo (VaR) de Acciones")
    themed_info(
        "El **Valor en Riesgo (VaR)** estima la pérdida máxima que un portafolio podría sufrir en un horizonte "
        "de tiempo (<span style='font-family: serif; font-style: italic;'>t</span>) con un nivel de confianza "
        "(<span style='font-family: serif; font-style: italic;'>1 − &alpha;</span>). <br><br>"
        "Calculamos el riesgo mediante el método paramétrico (asumiendo distribución normal) y mediante Simulación Monte Carlo "
        "(evaluando miles de escenarios históricos posibles para capturar caídas extremas)."
    )

    with st.expander("Seleccionar Activos y Periodo", expanded=True):
        st.markdown(
            "Selecciona **en qué empresas invertir** y **cuánto historial** se utilizará "
            "para estimar su volatilidad y rendimiento pasado."
        )
        c1, c2 = st.columns([2, 1])
        with c1:
            tickers_str = st.text_input(
                "Símbolos (separados por coma):", value="AAPL, MSFT, META",
                help="Ejemplos: AAPL, CEMEXCPO.MX, SPY, AMZN",
            )
        with c2:
            hoy = datetime.date.today()
            fecha_inicio = st.date_input(
                "Fecha de Inicio (Histórico)",
                value=hoy - datetime.timedelta(days=365 * 3),
            )

    tickers_list = [t.strip().upper() for t in tickers_str.split(",") if t.strip()]
    separador()

    st.markdown("#### Asignar Ponderaciones Manuales")
    st.caption("Haz clic en las celdas de la columna **Peso (%)** para ajustar qué porcentaje de tu capital va a cada acción.")

    if len(tickers_list) > 0:
        peso_eq = round(100.0 / len(tickers_list), 4)
        df_pesos_ini = pd.DataFrame({
            "Ticker":   tickers_list,
            "Peso (%)": [peso_eq] * len(tickers_list),
        })
        df_pesos_editado = st.data_editor(
            df_pesos_ini, hide_index=True, use_container_width=True,
            column_config={
                "Ticker":   st.column_config.TextColumn("Ticker", disabled=True),
                "Peso (%)": st.column_config.NumberColumn(
                    "Peso (%)", min_value=0.0, max_value=100.0,
                    step=0.01, format="%.4f%%"),
            },
        )
        suma_pesos = df_pesos_editado["Peso (%)"].sum()
        if abs(suma_pesos - 100.0) > 0.1:
            themed_warning(
                f"Tus porcentajes suman **{suma_pesos:.1f}%**. "
                "El modelo los ajustará matemáticamente para que sumen el 100% exacto."
            )
    else:
        themed_error("Ingresa al menos un ticker válido.")
        st.stop()

    separador()
    st.markdown("#### Configurar Escenario de Riesgo")
    st.markdown(
        "Define tu **inversión total**, qué tan conservador quieres ser en tu estimación (nivel de confianza) "
        "y a cuántos días hacia el futuro deseas proyectar la pérdida potencial."
    )

    col_r1, col_r2, col_r3 = st.columns(3)
    with col_r1:
        val_portafolio = st.number_input(
            "Capital Total Invertido ($)", min_value=100.0,
            value=100_000.0, step=10_000.0, key="riesgo_capital"
        )
    with col_r2:
        conf_str  = st.selectbox("Nivel de Confianza", ["95%", "99%"], index=1, key="riesgo_conf_sel")
        confianza = 0.95 if conf_str == "95%" else 0.99
    with col_r3:
        horizonte_str = st.selectbox(
            "Horizonte de Tiempo",
            ["1 Día", "10 Días", "21 Días (1 Mes)"],
            index=1,
        )
        dias_h = int(horizonte_str.split()[0])

    ejecutar_riesgo = st.button(
        "Calcular Métricas de Riesgo", use_container_width=True, key="btn_riesgo",
    )
    separador()

    if ejecutar_riesgo:
        dict_pesos = dict(zip(
            df_pesos_editado["Ticker"],
            df_pesos_editado["Peso (%)"] / 100.0,
        ))
        with st.spinner("Descargando precios y simulando Monte Carlo..."):
            try:
                data, rend_p, vol_p, pesos_reales, cols_reales = engine.evaluar_portafolio_personalizado(
                    tickers_list, dict_pesos, fecha_inicio, hoy
                )
                st.session_state.update({
                    "riesgo_data": data, "riesgo_rend": rend_p, "riesgo_vol": vol_p,
                    "riesgo_pesos": pesos_reales, "riesgo_cols": list(cols_reales),
                    "riesgo_capital_val": val_portafolio, "riesgo_confianza": confianza,
                    "riesgo_horizonte": horizonte_str, "riesgo_dias": dias_h,
                    "riesgo_hoy": hoy,
                })
                themed_success("Simulación de riesgo completada.")
            except Exception as e:
                themed_error(f"Error en el cálculo o descarga de datos: {e}")

    if "riesgo_rend" in st.session_state and "riesgo_cols" in st.session_state:
        rend_p       = st.session_state["riesgo_rend"]
        vol_p        = st.session_state["riesgo_vol"]
        pesos_reales = st.session_state["riesgo_pesos"]
        cols_reales  = st.session_state["riesgo_cols"]
        capital      = st.session_state["riesgo_capital_val"]
        conf         = st.session_state["riesgo_confianza"]
        h_str        = st.session_state["riesgo_horizonte"]
        dias         = st.session_state["riesgo_dias"]
        data         = st.session_state["riesgo_data"]
        hoy_g        = st.session_state["riesgo_hoy"]

        st.markdown("### Perfil del Portafolio")
        c_m1, c_m2, c_m3 = st.columns(3)
        c_m1.metric("Rendimiento Esperado Anual", f"{rend_p*100:.2f}%")
        c_m2.metric("Volatilidad Anual (σ)",       f"{vol_p*100:.2f}%")
        sharpe_ref = (rend_p - 0.05) / vol_p if vol_p > 0 else 0.0
        c_m3.metric("Ratio de Sharpe", f"{sharpe_ref:.4f}", help="Tasa libre de riesgo asumiendo 5% anual")
        separador()

        var_p,  _, _, _ = engine.calcular_var_parametrico(rend_p, vol_p, capital, conf, dias)
        var_mc, cvar_mc = engine.calcular_var_cvar_montecarlo(rend_p, vol_p, capital, conf, dias)

        st.markdown(f"### VaR — Horizonte: **{h_str}** | Confianza: **{conf*100:.0f}%**")
        col_res1, col_res2, col_res3 = st.columns(3)
        
        with col_res1:
            themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR Paramétrico: ${var_p:,.2f}</h4>")
        with col_res2:
            themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR Monte Carlo: ${var_mc:,.2f}</h4>")
        with col_res3:
            themed_error(f"<h4 style='margin:0; color:inherit;'>CVaR (Shortfall): ${cvar_mc:,.2f}</h4>")

        with paso_a_paso():
            st.latex(r"VaR_{param} = V_0 \times \left(Z_{\alpha} \times \sigma_{anual} \times \sqrt{\frac{t}{252}} - \mu_{anual} \times \frac{t}{252}\right)")
            z_score = norm.ppf(conf)
            t_frac = dias / 252
            st.latex(rf"VaR_{{param}} = {capital:,.2f} \left({z_score:.4f} \times {vol_p:.4f} \times \sqrt{{{t_frac:.4f}}} - {rend_p:.4f} \times {t_frac:.4f}\right)")
            factor_vol = z_score * vol_p * np.sqrt(t_frac)
            factor_ret = rend_p * t_frac
            st.latex(rf"VaR_{{param}} = {capital:,.2f} ({factor_vol:.6f} - {factor_ret:.6f}) = {var_p:,.2f}")
            st.write("---")
            st.latex(r"CVaR_{MC} = \mathbb{E}[L \mid L > VaR]")
            st.latex(rf"CVaR_{{MC}} = {cvar_mc:,.2f}")

        separador()

        tab_comp_p, tab_hist_p, tab_dl_p = st.tabs([
            "Composición Efectiva", "Precios Históricos", "Exportar Datos"
        ])
        with tab_comp_p:
            df_pie = pd.DataFrame({"Activo": cols_reales, "Peso": pesos_reales})
            fig_pie = px.pie(df_pie, values="Peso", names="Activo", hole=0.4,
                             color_discrete_sequence=plotly_colors())
            fig_pie.update_layout(height=380, **plotly_theme())
            st.plotly_chart(fig_pie, use_container_width=True)
        with tab_hist_p:
            precios_norm = (data / data.iloc[0]) * 100
            fig_hist = px.line(precios_norm, x=precios_norm.index, y=precios_norm.columns,
                               labels={"value": "Valor normalizado ($)", "Date": "Fecha"})
            fig_hist.update_layout(hovermode="x unified", **plotly_theme())
            st.plotly_chart(fig_hist, use_container_width=True)
        with tab_dl_p:
            st.download_button(
                "⬇️ Descargar Precios Históricos (.csv)",
                data=data.to_csv().encode("utf-8"),
                file_name=f"precios_{hoy_g}.csv", mime="text/csv",
                use_container_width=True,
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CREDITMETRICS
# ══════════════════════════════════════════════════════════════════════════════
with tab_cm:
    st.markdown("### CreditMetrics — Riesgo de Crédito en Portafolios de Bonos")
    themed_info(
        "**CreditMetrics** (J.P. Morgan, 1997) cuantifica el riesgo de crédito de un portafolio "
        "de bonos corporativos estimando todos sus valores posibles al final del año, "
        "basándose en la **matriz de transición** de calificaciones de S&P.<br><br>"
        "Soporta de **1 a 10 bonos**, evaluando el riesgo tanto de forma **independiente** (analítico exacto) "
        "como **correlacionado** (Cópula Gaussiana vía Monte Carlo)."
    )
    separador()

    # ──────────────────────────────────────────────────────────────────────────
    # CONSTANTES
    # ──────────────────────────────────────────────────────────────────────────
    CONF_LEVELS = [0.90, 0.95, 0.99, 0.999]
    CONF_LABELS = ["90%", "95%", "99%", "99.9%"]

    # ──────────────────────────────────────────────────────────────────────────
    # SUB-TABS
    # ──────────────────────────────────────────────────────────────────────────
    from credit_engine import NR_METHODS
    _NR_OPTS = ["raw_with_d", "redistribute"]
    col_opt1, col_opt2 = st.columns(2)
    with col_opt1:
        nr_mode = st.selectbox(
            "Tratamiento de NR (Not Rated):",
            _NR_OPTS,
            index=0,
            format_func=lambda x: NR_METHODS[x],
            key="cm_nr_mode",
            help="Controla cómo se manejan las probabilidades NR de la matriz S&P.",
        )
    with col_opt2:
        _mode_info = {
            "raw_with_d":   "Probabilidades S&P crudas originales. NR se excluye del sistema.",
            "redistribute": "Filas se normalizan a 1.0. NR redistribuido matemáticamente y de manera proporcional.",
        }
        themed_info(_mode_info[nr_mode])

    if (st.session_state.get("cm_last_mode") != nr_mode or "cm_tm" not in st.session_state):
        from credit_engine import _TM_RAW_17x19
        st.session_state["cm_tm"] = build_transition_matrix(_TM_RAW_17x19.copy(), nr_mode)
        st.session_state["cm_last_mode"] = nr_mode
        _raw_d = np.zeros((1, 19)); _raw_d[0, 17] = 1.0
        st.session_state["cm_tm_raw"] = np.vstack([_TM_RAW_17x19.copy(), _raw_d])

    _N_STATES = st.session_state["cm_tm"].shape[0]
    _RATINGS_WORK = RATINGS[:_N_STATES]
    _INC_D = (nr_mode != "raw_no_d_nr")   # True para raw_with_d, redistribute, simple_normalize
    separador()

    st1, st2, st3, st4, st5, st6 = st.tabs([
        "Bonos del Portafolio",
        "Matriz de Transición",
        "Curva de Tasas",
        "Caso Independiente",
        "Caso Correlacionado",
        "Exportar a Excel",
    ])

    # ── ST1: PARÁMETROS DE BONOS ──────────────────────────────────────────────
    with st1:
        st.markdown("#### Configurar bonos del portafolio (1 a 10 bonos)")
        themed_info(
            "Ingresa los parámetros financieros y contractuales de cada bono emitido. "
            "El sistema utilizará estos datos para revaluar el bono si la empresa cambia de calificación."
        )
        n_bonds = st.number_input("Número de bonos en el portafolio:", min_value=1, max_value=10,
                                   value=2, step=1, key="cm_n")
        _defs = [
            {"n":"Nvidia",  "r":"AA-", "vn":100., "c":5.0,  "t":5, "p":1, "rec":43.18},
            {"n":"Henkel",  "r":"A",   "vn":100., "c":5.0,  "t":5, "p":1, "rec":43.18},
            {"n":"Bono C",  "r":"BBB", "vn":100., "c":6.0,  "t":3, "p":2, "rec":40.0},
            {"n":"Bono D",  "r":"BB",  "vn":100., "c":7.0,  "t":3, "p":2, "rec":35.0},
            {"n":"Bono E",  "r":"B+",  "vn":100., "c":8.0,  "t":3, "p":2, "rec":30.0},
            {"n":"Bono F",  "r":"BBB+","vn":100., "c":5.5,  "t":4, "p":2, "rec":40.0},
            {"n":"Bono G",  "r":"A+",  "vn":100., "c":4.5,  "t":4, "p":1, "rec":43.0},
            {"n":"Bono H",  "r":"AA",  "vn":100., "c":4.0,  "t":3, "p":1, "rec":43.0},
            {"n":"Bono I",  "r":"BBB-","vn":100., "c":6.5,  "t":5, "p":2, "rec":40.0},
            {"n":"Bono J",  "r":"B",   "vn":100., "c":9.0,  "t":3, "p":4, "rec":35.0},
        ]
        bond_params = []
        for i in range(int(n_bonds)):
            d = _defs[i]
            st.markdown(f"##### Bono {i+1}")
            co1, co2, co3, co4 = st.columns(4)
            with co1:
                nom = st.text_input("Nombre de Emisora", value=d["n"], key=f"cm_nom_{i}")
                rat = st.selectbox("Calificación S&P Actual", RATINGS[:-1],
                                   index=RATINGS.index(d["r"]) if d["r"] in RATINGS else 0,
                                   key=f"cm_r_{i}")
            with co2:
                vn = st.number_input("Nominal ($)", min_value=1., value=d["vn"],
                                      step=10., key=f"cm_vn_{i}")
                cp = st.number_input("Cupón anual (%)", min_value=0., value=d["c"],
                                      step=0.25, key=f"cm_c_{i}") / 100
            with co3:
                T  = st.number_input("Vencimiento (años)", min_value=1, max_value=10,
                                      value=d["t"], step=1, key=f"cm_T_{i}")
                pg = st.selectbox("Pagos/año", [1,2,4,12],
                                   index=[1,2,4,12].index(d["p"]), key=f"cm_p_{i}",
                                   format_func=lambda x: {1:"Anual",2:"Semestral",
                                                           4:"Trimestral",12:"Mensual"}[x])
            with co4:
                rc = st.number_input("Recuperación D (%)", min_value=0., max_value=100.,
                                      value=d["rec"], step=1., key=f"cm_rc_{i}",
                                      help="Porcentaje que se recupera si el bono cae en Default (D).") / 100
            bond_params.append({
                "nombre": nom, "rating": rat, "rating_idx": RATING_IDX[rat],
                "VN": vn, "cupon_pct": cp, "T": int(T), "pagos": pg, "recov": rc,
            })
            if i < int(n_bonds) - 1: st.markdown("---")
        st.session_state["cm_bparams"] = bond_params
        themed_success(f"**{int(n_bonds)} bono(s)** cargados en la memoria del modelo.")

    # ── ST2: MATRIZ DE TRANSICIÓN ─────────────────────────────────────────────
    with st2:
        st.markdown("#### Matriz de Transición de Calificaciones S&P")
        themed_info(
            "Muestra las **probabilidades históricas** de migración de calificación en un año (S&P Global, 1981–2021). "
            "Las filas son la calificación actual del emisor y las columnas la calificación destino."
        )

        if "cm_tm_raw" not in st.session_state:
            from credit_engine import _TM_RAW_17x19, RATINGS_EMIT as _RE
            _raw = _TM_RAW_17x19.copy()
            _d_row = np.zeros((1, 19)); _d_row[0, 17] = 1.0
            st.session_state["cm_tm_raw"] = np.vstack([_raw, _d_row])
        if "cm_tm" not in st.session_state:
            st.session_state["cm_tm"] = DEFAULT_TM.copy()

        col_tm1, col_tm2 = st.columns([4, 1])
        with col_tm2:
            if st.button("Restaurar S&P Original", key="cm_rst_tm"):
                from credit_engine import _TM_RAW_17x19
                _raw = _TM_RAW_17x19.copy()
                _d = np.zeros((1,19)); _d[0,17] = 1.0
                st.session_state["cm_tm_raw"] = np.vstack([_raw, _d])
                st.session_state["cm_tm"] = DEFAULT_TM.copy()
                st.rerun()
            row_sums = st.session_state["cm_tm_raw"].sum(axis=1)
            df_sums = pd.DataFrame({
                "Rating": RATINGS,
                "Suma": [f"{s:.4f}" for s in row_sums],
                "": ["✓" if abs(s-1)<0.002 else "⚠" for s in row_sums]
            })
            st.caption("Verificación de filas (suman 1):")
            st.dataframe(df_sums, hide_index=True, use_container_width=True, height=460)

        with col_tm1:
            from credit_engine import RATINGS_EMIT as _RALL
            _DEST_LABELS = _RALL[:17] + ["D", "NR"]
            df_tm_edit = pd.DataFrame(
                st.session_state["cm_tm_raw"],
                index=RATINGS,
                columns=_DEST_LABELS,
            )
            df_tm_edit.index.name = "From / To"
            ed = st.data_editor(
                df_tm_edit.round(6), use_container_width=True, height=570, num_rows="fixed",
                column_config={c: st.column_config.NumberColumn(
                    c, format="%.4f", step=0.0001, min_value=0., max_value=1.)
                    for c in _DEST_LABELS},
            )
            raw_arr = ed.values.astype(float)
            st.session_state["cm_tm_raw"] = raw_arr
            _active_mode = st.session_state.get("cm_nr_mode", "redistribute")
            st.session_state["cm_tm"] = build_transition_matrix(raw_arr[:17], _active_mode)

        c_th = get_current_theme()
        _tm_plot = st.session_state.get('cm_tm', DEFAULT_TM)
        fig_hm = go.Figure(go.Heatmap(
            z=_tm_plot[:min(_N_STATES, N_R-1), :min(_N_STATES, N_R-1)],
            x=list(_RATINGS_WORK[:min(_N_STATES, N_R-1)]),
            y=list(_RATINGS_WORK[:min(_N_STATES, N_R-1)]),
            colorscale=[[0,"#FFFFFF"],[1, c_th["primary"]]],
            text=np.round(_tm_plot[:min(_N_STATES, N_R-1), :min(_N_STATES, N_R-1)]*100, 3),
            texttemplate="%{text:.2f}%", showscale=True,
        ))
        fig_hm.update_layout(
            title="Mapa de calor — Probabilidades de transición (%)",
            xaxis_title="Rating destino", yaxis_title="Rating origen",
            height=550, **plotly_theme(),
        )
        st.plotly_chart(fig_hm, use_container_width=True)

        with paso_a_paso():
            if nr_mode == "raw_no_d_nr":
                st.latex(r"\text{Suma de fila no controlada (sin normalizar)}")
            else:
                st.latex(r"P_{i,j}^{adj} = \frac{P_{i,j}}{\sum_{k=1}^{N} P_{i,k}}")
                st.latex(rf"\sum P_{{AAA}} = {np.sum(_tm_plot[0]):.6f}")

    # ── HELPERS: curva de tasas automática ───────────────────────────────────
    _DEFAULT_SPREADS_FLAT = DEFAULT_SPREADS[:17, 0] - DEFAULT_TREASURY[0]

    def _default_tsy_anchors(max_t: int) -> np.ndarray:
        v = np.full(max_t, DEFAULT_TREASURY[-1])
        v[:min(max_t, len(DEFAULT_TREASURY))] = DEFAULT_TREASURY[:min(max_t, len(DEFAULT_TREASURY))]
        return v

    def _build_allin_table(max_t: int) -> tuple:
        tsy  = st.session_state.get("cm_tsy_anchors", _default_tsy_anchors(max_t))
        spr  = st.session_state.get("cm_spreads_flat", _DEFAULT_SPREADS_FLAT.copy())
        if len(tsy) < max_t:
            tsy = np.append(tsy, np.full(max_t - len(tsy), tsy[-1]))
        tenors = np.arange(0.5, max_t + 0.01, 0.5)
        anchor_x = np.array([0] + list(range(1, max_t + 1)))
        allin = np.zeros((17, len(tenors)))
        for r_idx in range(17):
            anchor_y = np.array([0.0] + [tsy[y] + spr[r_idx] for y in range(max_t)])
            allin[r_idx] = np.interp(tenors, anchor_x, anchor_y)
        full = np.vstack([allin, np.full((1, len(tenors)), np.nan)])
        return full, tenors

    # ── ST3: CURVA DE TASAS ───────────────────────────────────────────────────
    with st3:
        st.markdown("#### Curvas de Rendimiento (Treasury + Spread)")
        themed_info(
            "Ingresa la **tasa libre de riesgo** (Treasury) y el **spread crediticio** por calificación. "
            "El sistema calculará automáticamente la tasa 'All-In' que se usará para descontar los flujos futuros de los bonos "
            "en cada escenario posible."
        )

        _bond_ps = st.session_state.get("cm_bparams", [])
        _max_T   = min(max((b["T"] for b in _bond_ps), default=5), 10)
        _n_spr_rows = min(17, _N_STATES)

        if "cm_tsy_anchors" not in st.session_state:
            st.session_state["cm_tsy_anchors"] = _default_tsy_anchors(_max_T)
        else:
            old_t = st.session_state["cm_tsy_anchors"]
            if len(old_t) != _max_T:
                new_t = _default_tsy_anchors(_max_T)
                new_t[:min(_max_T, len(old_t))] = old_t[:min(_max_T, len(old_t))]
                st.session_state["cm_tsy_anchors"] = new_t

        if "cm_spreads_flat" not in st.session_state:
            st.session_state["cm_spreads_flat"] = _DEFAULT_SPREADS_FLAT.copy()

        col3a, col3b = st.columns([1, 2])

        with col3a:
            st.markdown("##### Tasa libre de riesgo (rf)")
            df_tsy = pd.DataFrame({
                "Año": list(range(1, _max_T + 1)),
                "Tasa (%)": (st.session_state["cm_tsy_anchors"] * 100).round(4),
            })
            ed_tsy = st.data_editor(
                df_tsy, hide_index=True, use_container_width=True,
                column_config={
                    "Año":      st.column_config.NumberColumn(disabled=True),
                    "Tasa (%)": st.column_config.NumberColumn(
                        format="%.4f%%", step=0.01, min_value=0.0),
                },
            )
            st.session_state["cm_tsy_anchors"] = ed_tsy["Tasa (%)"].values / 100

            if st.button("Restaurar tasas de mercado", key="cm_rst_rates"):
                st.session_state.pop("cm_tsy_anchors", None)
                st.session_state.pop("cm_spreads_flat", None)
                st.rerun()

        with col3b:
            st.markdown("##### Spread crediticio")
            df_spr = pd.DataFrame({
                "Rating":     RATINGS[:_n_spr_rows],
                "Spread (%)": (st.session_state["cm_spreads_flat"][:_n_spr_rows] * 100).round(4),
            })
            ed_spr = st.data_editor(
                df_spr, hide_index=True, use_container_width=True,
                height=min(38 + _n_spr_rows * 35, 630),
                column_config={
                    "Rating":     st.column_config.TextColumn(disabled=True),
                    "Spread (%)": st.column_config.NumberColumn(
                        format="%.4f%%", step=0.01, min_value=0.0),
                },
            )
            new_spr = st.session_state["cm_spreads_flat"].copy()
            new_spr[:_n_spr_rows] = ed_spr["Spread (%)"].values / 100
            st.session_state["cm_spreads_flat"] = new_spr

        separador()
        allin_preview, tenors_preview = _build_allin_table(_max_T)
        tenor_labels = [f"t={t:.1f}" for t in tenors_preview]

        with st.expander("Ver tabla All-In generada (Tasa rf + Spread)", expanded=False):
            df_preview = pd.DataFrame(
                allin_preview[:_n_spr_rows] * 100,
                index=RATINGS[:_n_spr_rows],
                columns=tenor_labels,
            ).reset_index()
            st.dataframe(
                df_preview,
                use_container_width=True,
                hide_index=True,
                height=min(38 + _n_spr_rows * 35, 560),
                column_config={
                    "Rating": st.column_config.TextColumn(),
                    **{c: st.column_config.NumberColumn(format="%.4f%%") for c in tenor_labels}
                },
            )

        c_th = get_current_theme()
        fig_yc = go.Figure()
        tsy_line = np.interp(
            tenors_preview,
            [0] + list(range(1, _max_T + 1)),
            [0] + list(st.session_state["cm_tsy_anchors"])
        )
        fig_yc.add_trace(go.Scatter(
            x=tenors_preview.tolist(), y=(tsy_line * 100).tolist(),
            name="Tesoro (rf)", mode="lines+markers",
            line=dict(color=c_th["primary"], width=3, dash="dash"),
        ))
        for r_name, r_row in zip(RATINGS[:_n_spr_rows], allin_preview[:_n_spr_rows]):
            if r_name in ["AAA", "AA", "A", "BBB", "BB", "B", "CCC/C"]:
                fig_yc.add_trace(go.Scatter(
                    x=tenors_preview.tolist(), y=(r_row * 100).tolist(),
                    name=r_name, mode="lines", opacity=0.8,
                ))
        fig_yc.update_layout(
            title="Curvas de rendimiento All-In por calificación",
            xaxis_title="Plazo (años)", yaxis_title="Tasa all-in (%)",
            height=400, **plotly_theme(),
        )
        st.plotly_chart(fig_yc, use_container_width=True)

    # ── HELPER: datos de bonos con valores calculados ─────────────────────────
    def _get_bond_data():
        params = st.session_state.get("cm_bparams", [])
        mode   = st.session_state.get("cm_nr_mode", "redistribute")
        inc_d  = (mode != "raw_no_d_nr")
        if not params:
            return []
        max_t  = max(b["T"] for b in params)
        allin, tenors = _build_allin_table(max_t)
        out = []
        for bp in params:
            vals = bond_values_per_rating(
                bp["VN"], bp["cupon_pct"], bp["T"],
                bp["pagos"], bp["recov"], allin,
                include_d=inc_d,
                spread_times=tenors,
            )
            out.append({**bp, "values": vals})
        return out

    def _build_metrics_table(scaled: dict, conf_levels, conf_labels, ev: float) -> pd.DataFrame:
        rows = []
        for cf, lb in zip(conf_levels, conf_labels):
            r = scaled[cf]
            rows.append({
                "Confianza":         lb,
                "VaR 1 año ($)":     f"${r['VaR_1y']:,.4f}",
                "CVaR 1 año ($)":    f"${r['CVaR_1y']:,.4f}",
                "VaR 1 día ($)":     f"${r['VaR_1d']:,.4f}",
                "CVaR 1 día ($)":    f"${r['CVaR_1d']:,.4f}",
                "VaR 10 días ($)":   f"${r['VaR_10d']:,.4f}",
                "CVaR 10 días ($)":  f"${r['CVaR_10d']:,.4f}",
                "Capital (3×VaR10d)":f"${r['Capital']:,.4f}",
                "VaR 1y % E[V]":     f"{r['VaR_1y']/ev*100:.2f}%" if ev > 0 else "—",
            })
        return pd.DataFrame(rows)

    # ── ST4: CASO INDEPENDIENTE ───────────────────────────────────────────────
    with st4:
        st.markdown("#### Cálculo VaR: Caso Independiente")
        themed_info(
            "Bajo el supuesto de independencia, calculamos todos los escenarios futuros posibles usando la "
            "**convolución exacta**. Esto genera la distribución probabilística perfecta de tu portafolio si "
            "el impago de una empresa no contagiara a las demás."
        )

        if st.button("Calcular distribución analítica", use_container_width=True, key="btn_ind"):
            with st.spinner("Calculando convolución..."):
                bd = _get_bond_data()
                if not bd:
                    themed_error("Configura los bonos primero.")
                else:
                    tm      = st.session_state.get("cm_tm", DEFAULT_TM)
                    mode_now = st.session_state.get("cm_nr_mode", "redistribute")
                    do_normalize = (mode_now != "raw_with_d")

                    sdist  = independent_distribution(bd, tm)
                    vr     = var_cvar_from_distribution(sdist, CONF_LEVELS,
                                                        normalize=do_normalize)
                    scaled = scale_var_cvar(vr, CONF_LEVELS)

                    ev_ex    = vr[CONF_LEVELS[0]]["EV"]
                    sigma_ex = vr[CONF_LEVELS[0]]["sigma"]
                    vr_p     = var_cvar_parametric(ev_ex, sigma_ex, CONF_LEVELS)
                    scaled_p = scale_var_cvar(vr_p, CONF_LEVELS)

                    st.session_state.update({
                        "cm_sdist":        sdist,
                        "cm_ivars":        vr,
                        "cm_iscaled":      scaled,
                        "cm_ivars_param":  vr_p,
                        "cm_iscaled_param":scaled_p,
                        "cm_ibonds":       bd,
                        "cm_ev":           vr[0.99]["EV"],
                    })
                    themed_success(f"La máquina probó matemáticamente **{len(sdist):,}** escenarios distintos.")

        if "cm_ivars_param" in st.session_state:
            bd_i  = st.session_state["cm_ibonds"]
            sdist = st.session_state["cm_sdist"]
            vr    = st.session_state["cm_ivars"]
            sc_p  = st.session_state["cm_iscaled_param"]
            ev    = vr[CONF_LEVELS[0]]["EV"]
            sigma = vr[CONF_LEVELS[0]]["sigma"]
            c_th  = get_current_theme()

            col_ev1, col_ev2 = st.columns(2)
            col_ev1.metric("Valor Esperado (E[V])", f"${ev:,.4f}")
            col_ev2.metric("Volatilidad (σ)",    f"${sigma:,.4f}")
            separador()

            st.markdown("##### Requerimientos Regulatorios y Extremos de Riesgo")
            themed_info(
                f"La regulación bancaria actual solicita escalar el VaR anual a ventanas más cortas usando "
                f"la raíz cuadrada del tiempo (<span style='font-family: serif; font-style: italic;'>&radic;t</span>). "
                f"El requerimiento de Capital dictado por Basilea asume un colchón de 3 veces el VaR a 10 días."
            )
            df_metrics_p = _build_metrics_table(sc_p, CONF_LEVELS, CONF_LABELS, ev)
            st.dataframe(df_metrics_p, hide_index=True, use_container_width=True)

            r99p = sc_p[0.99]
            col_p1, col_p2, col_p3, col_p4 = st.columns(4)
            with col_p1: themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR 1d (99%):<br>${r99p['VaR_1d']:,.2f}</h4>")
            with col_p2: themed_warning(f"<h4 style='margin:0; color:inherit;'>CVaR 1d (99%):<br>${r99p['CVaR_1d']:,.2f}</h4>")
            with col_p3: themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR 10d (99%):<br>${r99p['VaR_10d']:,.2f}</h4>")
            with col_p4: themed_error(f"<h4 style='margin:0; color:inherit;'>Capital Basilea:<br>${r99p['Capital']:,.2f}</h4>")

            with paso_a_paso():
                st.latex(r"E[V]_{port} = \sum_{i=1}^n \sum_{j=1}^{K} p_{i,j} \times V_{i,j}")
                st.latex(r"VaR_{1d} = \frac{VaR_{1y}}{\sqrt{252}}")
                st.latex(r"VaR_{10d} = VaR_{1d} \times \sqrt{10}")
                st.latex(r"Capital = 3 \times VaR_{10d}")

            separador()

            st.markdown("##### Distribución individual por bono")
            tm4 = st.session_state.get("cm_tm", DEFAULT_TM)
            tabs_bonds = st.tabs([b["nombre"] for b in bd_i])
            for tab_b, b in zip(tabs_bonds, bd_i):
                with tab_b:
                    probs_b = tm4[b["rating_idx"]]
                    col4a, col4b = st.columns([1, 2])
                    with col4a:
                        _n_states_b = len(b["values"])
                        _labels_b = list(_RATINGS_WORK[:_n_states_b])
                        df_bv = pd.DataFrame({
                            "Rating destino": _labels_b,
                            "Probabilidad":   [f"{p:.4%}" for p in probs_b[:_n_states_b]],
                            "Valor ($)":      [f"${v:.4f}" for v in b["values"]],
                            "VxP ($)":        [f"${v*p:.6f}" for v,p in zip(b["values"],probs_b[:_n_states_b])],
                        })
                        st.dataframe(df_bv, hide_index=True, use_container_width=True, height=430)
                        ev_b = sum(v*p for v,p in zip(b["values"],probs_b[:_n_states_b]))
                        var_b= sum((v-ev_b)**2*p for v,p in zip(b["values"],probs_b[:_n_states_b]))
                        st.metric("E[V] bono", f"${ev_b:.4f}")
                        st.metric("sigma bono", f"${var_b**0.5:.4f}")
                    with col4b:
                        fig_bv = go.Figure()
                        colors_bv = [c_th["accent"] if r==b["rating"] else c_th["primary"]
                                     for r in _labels_b]
                        fig_bv.add_trace(go.Bar(
                            x=_labels_b, y=probs_b[:_n_states_b]*100,
                            text=[f"${v:.2f}" for v in b["values"]],
                            textposition="outside",
                            marker_color=colors_bv,
                            name="Prob (%)",
                        ))
                        fig_bv.add_hline(y=0, line_color=c_th["border"])
                        fig_bv.update_layout(
                            title=f"{b['nombre']} — Rating actual: {b['rating']}",
                            xaxis_title="Rating destino", yaxis_title="Probabilidad (%)",
                            height=400, **plotly_theme(),
                        )
                        st.plotly_chart(fig_bv, use_container_width=True)

            separador()

            st.markdown("##### Probabilidad Combinada (Portafolio Total)")
            vr_p  = st.session_state["cm_ivars_param"]
            vals_a  = np.array([d[0] for d in sdist])
            probs_a = np.array([d[1] for d in sdist]); probs_a /= probs_a.sum()
            cum_a   = np.cumsum(probs_a)

            fig_d = go.Figure()
            fig_d.add_trace(go.Bar(
                x=vals_a, y=probs_a*100, name="P(V=v) (%)",
                marker_color=c_th["primary"], opacity=0.65,
            ))
            fig_d.add_trace(go.Scatter(
                x=vals_a, y=cum_a*100, name="CDF (%)",
                line=dict(color=c_th["accent"], width=2), yaxis="y2",
            ))
            for cf, lb in zip(CONF_LEVELS, CONF_LABELS):
                q_param = ev - vr_p[cf]["VaR"]   # cuantil paramétrico
                fig_d.add_vline(
                    x=q_param, line_dash="dot", line_color=c_th["danger"],
                    annotation_text=f"VaR {lb}", annotation_position="top",
                )
            fig_d.add_vline(x=ev, line_color=c_th["success"],
                            annotation_text=f"E[V]={ev:.2f}")
            fig_d.update_layout(
                title="Distribución del valor del portafolio",
                xaxis_title="Valor ($)", yaxis_title="Probabilidad (%)",
                yaxis2=dict(title="CDF (%)", overlaying="y", side="right"),
                height=440, **plotly_theme(),
            )
            st.plotly_chart(fig_d, use_container_width=True)

            st.markdown("##### Tabla de distribución completa")
            df_fd = pd.DataFrame({
                "Valor ($)":           vals_a,
                "Probabilidad (%)":    probs_a * 100,
                "CDF (%)":             cum_a * 100,
                "Pérdida vs E[V] ($)": ev - vals_a,
            }).sort_values("Valor ($)", ascending=False).reset_index(drop=True)

            _MAX_ROWS_DISPLAY = 5_000
            df_show = df_fd.head(_MAX_ROWS_DISPLAY)
            if len(df_fd) > _MAX_ROWS_DISPLAY:
                themed_info(
                    f"Mostrando las primeras **{_MAX_ROWS_DISPLAY:,}** filas de "
                    f"**{len(df_fd):,}** escenarios. "
                    "Descarga el Excel para ver la distribución completa."
                )
            st.dataframe(
                df_show,
                use_container_width=True,
                height=400,
                hide_index=True,
                column_config={
                    "Valor ($)":           st.column_config.NumberColumn(format="$%.4f"),
                    "Probabilidad (%)":    st.column_config.NumberColumn(format="%.6f%%"),
                    "CDF (%)":             st.column_config.NumberColumn(format="%.4f%%"),
                    "Pérdida vs E[V] ($)": st.column_config.NumberColumn(format="$%.4f"),
                },
            )
            st.session_state["cm_df_dist"] = df_fd

    # ── ST5: CASO CORRELACIONADO ──────────────────────────────────────────────
    with st5:
        st.markdown("#### Cálculo VaR: Caso Correlacionado (Cópula Gaussiana)")
        themed_info(
            "En las crisis globales, todas las empresas caen juntas. La **Cópula Gaussiana** es una herramienta que conecta las "
            "probabilidades individuales de los bonos simulando su correlación en la bolsa de valores. Si una empresa se degrada, "
            "aumenta la probabilidad de que la otra también lo haga."
        )

        bd_c = _get_bond_data()
        n_bc = len(bd_c)

        col5a, col5b = st.columns([2, 1])
        with col5a:
            st.markdown("##### Matriz de correlación entre empresas")
            st.caption("Usa la relación que tienen sus acciones en la bolsa como 'proxy' de correlación.")
            if ("cm_corrm" not in st.session_state or
                    st.session_state["cm_corrm"].shape[0] != n_bc):
                st.session_state["cm_corrm"] = np.eye(n_bc)
            noms_bc = [b["nombre"] for b in bd_c]
            df_cr = pd.DataFrame(st.session_state["cm_corrm"], index=noms_bc, columns=noms_bc)
            df_cr.index.name = "Bono"
            ed_cr = st.data_editor(
                df_cr.round(4), use_container_width=True,
                column_config={c: st.column_config.NumberColumn(
                    c, format="%.4f", step=0.01, min_value=-1., max_value=1.)
                    for c in noms_bc},
            )
            cm_arr = ed_cr.values.astype(float)
            np.fill_diagonal(cm_arr, 1.0)
            st.session_state["cm_corrm"] = cm_arr

        with col5b:
            n_sims5 = st.select_slider(
                "Muestras Monte Carlo:", options=[10_000,50_000,100_000,200_000],
                value=50_000, key="cm_sims5",
            )

            st.markdown("##### Umbrales Límite")
            tm5 = st.session_state.get("cm_tm", DEFAULT_TM)
            th_rows = []
            for b in bd_c:
                _n_th = st.session_state["cm_tm"].shape[0]
                cum_p = np.cumsum(tm5[b["rating_idx"]])
                with np.errstate(all='ignore'):
                    thresh = norm.ppf(np.clip(cum_p, 1e-15, 1-1e-15))
                thresh[-1] = np.inf
                th_rows.append([b["nombre"], b["rating"]] +
                               [f"{t:.3f}" if np.isfinite(t) else "∞" for t in thresh[:8]])
            df_th = pd.DataFrame(th_rows,
                                 columns=["Bono","Rating"]+[f"z({r})" for r in list(_RATINGS_WORK[:8])])
            st.dataframe(df_th, hide_index=True, use_container_width=True, height=280)

        separador()

        if st.button("Ejecutar Simulación de Contagio", use_container_width=True, key="btn_corr5"):
            with st.spinner(f"Simulando {n_sims5:,} caminos usando la matriz Cholesky..."):
                sims5 = gaussian_copula_simulation(
                    bd_c, st.session_state.get("cm_tm", DEFAULT_TM),
                    st.session_state["cm_corrm"],
                    n_sims=n_sims5, seed=None,
                )
                vc5_raw = var_cvar_from_simulations(sims5, CONF_LEVELS)
                ev5_c   = vc5_raw[CONF_LEVELS[0]]["EV"]
                sg5_c   = vc5_raw[CONF_LEVELS[0]]["sigma"]
                vc5     = var_cvar_parametric(ev5_c, sg5_c, CONF_LEVELS)
                sc5     = scale_var_cvar(vc5, CONF_LEVELS)
                st.session_state["cm_csims"]   = sims5
                st.session_state["cm_cvars"]   = vc5
                st.session_state["cm_cscaled"] = sc5
            themed_success(f"Simulación masiva completada.")

        if "cm_cvars" in st.session_state:
            vc5  = st.session_state["cm_cvars"]
            sc5  = st.session_state.get("cm_cscaled", scale_var_cvar(vc5, CONF_LEVELS))
            s5   = st.session_state["cm_csims"]
            ev5  = vc5[0.95]["EV"]
            sg5  = vc5[0.95]["sigma"]
            c_th = get_current_theme()

            col_ev5a, col_ev5b = st.columns(2)
            col_ev5a.metric("E[V] simulado",  f"${ev5:,.4f}")
            col_ev5b.metric("σ simulado",     f"${sg5:,.4f}")
            
            with paso_a_paso():
                st.latex(r"Z_i = \sqrt{\rho} M + \sqrt{1-\rho} \epsilon_i")
                st.latex(r"\text{Threshold}_j = \Phi^{-1}\left(\sum_{k=1}^j p_k\right)")

            st.markdown("##### Métricas de Riesgo y Contagio")
            df_metrics5 = _build_metrics_table(sc5, CONF_LEVELS, CONF_LABELS, ev5)
            st.dataframe(df_metrics5, hide_index=True, use_container_width=True)

            r99c = sc5[0.99]
            col_c1, col_c2, col_c3, col_c4 = st.columns(4)
            with col_c1: themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR 1d (99%):<br>${r99c['VaR_1d']:,.2f}</h4>")
            with col_c2: themed_warning(f"<h4 style='margin:0; color:inherit;'>CVaR 1d (99%):<br>${r99c['CVaR_1d']:,.2f}</h4>")
            with col_c3: themed_warning(f"<h4 style='margin:0; color:inherit;'>VaR 10d (99%):<br>${r99c['VaR_10d']:,.2f}</h4>")
            with col_c4: themed_error(f"<h4 style='margin:0; color:inherit;'>Capital Basilea:<br>${r99c['Capital']:,.2f}</h4>")
            separador()

            if "cm_iscaled_param" in st.session_state:
                st.markdown("##### La Prueba de Fuego: Riesgo Cíclico vs Aislado")
                themed_info(
                    "Si los activos de una cartera están fuertemente correlacionados, cuando el mercado colapsa, "
                    "caen juntos. Esta tabla te demuestra cómo el VaR sube considerablemente cuando activas "
                    "las simulaciones de correlación."
                )
                vi_sc = st.session_state["cm_iscaled_param"]
                comp = []
                for cf, lb in zip(CONF_LEVELS, CONF_LABELS):
                    ri = vi_sc[cf]; rc5 = sc5[cf]
                    delta_10d = rc5["VaR_10d"] - ri["VaR_10d"]
                    comp.append({
                        "Confianza":               lb,
                        "VaR 1d Indep. ($)":       f"${ri['VaR_1d']:,.4f}",
                        "VaR 1d Corr. ($)":        f"${rc5['VaR_1d']:,.4f}",
                        "VaR 10d Indep. ($)":      f"${ri['VaR_10d']:,.4f}",
                        "VaR 10d Corr. ($)":       f"${rc5['VaR_10d']:,.4f}",
                        "Capital Indep. ($)":      f"${ri['Capital']:,.4f}",
                        "Capital Corr. ($)":       f"${rc5['Capital']:,.4f}",
                        "Δ VaR 10d ($)":           f"${delta_10d:+,.4f}",
                        "Efecto correlación":      "Peligro Sistémico" if delta_10d > 0 else "Estabilidad",
                    })
                st.dataframe(pd.DataFrame(comp), hide_index=True, use_container_width=True)

            fig5 = go.Figure()
            fig5.add_trace(go.Histogram(
                x=s5, nbinsx=120, name="Distribución simulada",
                marker_color=c_th["primary"], opacity=0.7,
                histnorm="probability density",
            ))
            fig5.add_vline(x=ev5, line_color=c_th["success"], line_width=2,
                           annotation_text=f"E[V]={ev5:.2f}")
            for cf, lb in zip([0.95, 0.99, 0.999], ["95%","99%","99.9%"]):
                q_param = ev5 - vc5[cf]["VaR"]   # cuantil paramétrico
                fig5.add_vline(x=q_param, line_dash="dot", line_color=c_th["danger"],
                               annotation_text=f"VaR {lb}")
            fig5.update_layout(
                title=f"Histograma de Cópula Gaussiana ({n_sims5:,} escenarios)",
                xaxis_title="Valor del portafolio ($)", yaxis_title="Densidad Probabilística",
                height=420, **plotly_theme(),
            )
            st.plotly_chart(fig5, use_container_width=True)

    # ── ST6: EXPORTAR A EXCEL ─────────────────────────────────────────────────
    with st6:
        st.markdown("#### Exportar Arquitectura a Excel")
        themed_info(
            "Extrae la memoria completa de las simulaciones y parámetros a una hoja de Excel nativa. "
            "Perfecto para auditar, presentar o continuar trabajando los datos sin Python."
        )

        if st.button("Generar Excel de CreditMetrics", use_container_width=True, key="btn_xls"):
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                HDR  = Font(bold=True, color="FFFFFF", size=10)
                FIL1 = PatternFill("solid", start_color="203F9A")
                FIL2 = PatternFill("solid", start_color="E84797")
                FIL3 = PatternFill("solid", start_color="F2C8D8")
                CTR  = Alignment(horizontal="center")

                def hdr(ws, r, c, v, fill=FIL1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.font = HDR; cell.fill = fill; cell.alignment = CTR

                # Sheet 1: Bond params
                ws1 = wb.active; ws1.title = "1 Parametros"
                for j, h in enumerate(["Nombre","Rating","VN","Cupon%","T","Pagos/año","Recup%"], 1):
                    hdr(ws1, 1, j, h); ws1.column_dimensions[get_column_letter(j)].width = 14
                for i, b in enumerate(st.session_state.get("cm_bparams",[]), 2):
                    for j, v in enumerate([b["nombre"],b["rating"],b["VN"],
                                           b["cupon_pct"]*100,b["T"],b["pagos"],b["recov"]*100], 1):
                        ws1.cell(row=i, column=j, value=v).fill = FIL3

                # Sheet 2: Transition matrix
                ws2 = wb.create_sheet("2 Matriz Transicion")
                hdr(ws2, 1, 1, "From\\To", FIL2)
                tm_x = st.session_state.get("cm_tm", DEFAULT_TM)
                for j, r in enumerate(RATINGS, 2):
                    hdr(ws2, 1, j, r); ws2.column_dimensions[get_column_letter(j)].width = 8
                ws2.column_dimensions["A"].width = 8
                for i, rf in enumerate(RATINGS):
                    ws2.cell(row=i+2, column=1, value=rf).font = Font(bold=True)
                    for j in range(N_R):
                        c = ws2.cell(row=i+2, column=j+2, value=float(tm_x[i,j]))
                        c.number_format = "0.0000%"; c.alignment = CTR
                        if i == j: c.fill = FIL3

                # Sheet 3: Rates
                ws3 = wb.create_sheet("3 Tasas")
                _bp_x   = st.session_state.get("cm_bparams", [])
                _maxT_x = min(max((b["T"] for b in _bp_x), default=5), 10) if _bp_x else 5
                tsy_x   = st.session_state.get("cm_tsy_anchors", _default_tsy_anchors(_maxT_x))
                spr_x   = st.session_state.get("cm_spreads_flat", _DEFAULT_SPREADS_FLAT.copy())
                allin_x, tenors_x = _build_allin_table(_maxT_x)

                # Treasury row
                hdr(ws3, 1, 1, "Tasa libre de riesgo (tesoro)", FIL2)
                for j, yr in enumerate(range(1, _maxT_x + 1), 2):
                    hdr(ws3, 1, j, f"Año {yr}")
                    ws3.column_dimensions[get_column_letter(j)].width = 11
                ws3.column_dimensions["A"].width = 28
                for j, v in enumerate(tsy_x, 2):
                    ws3.cell(row=2, column=j, value=float(v)).number_format = "0.0000%"

                # Flat spreads
                hdr(ws3, 4, 1, "Spread crediticio (plano por rating)", FIL2)
                hdr(ws3, 4, 2, "Spread (%)")
                ws3.column_dimensions["B"].width = 14
                for i, (rn, sv) in enumerate(zip(RATINGS[:_n_spr_rows], spr_x[:_n_spr_rows]), 5):
                    ws3.cell(row=i, column=1, value=rn).font = Font(bold=True)
                    ws3.cell(row=i, column=2, value=float(sv)).number_format = "0.0000%"

                # All-in table (generated)
                _t_start_row = 5 + _n_spr_rows + 1
                hdr(ws3, _t_start_row, 1, "Tabla all-in generada (tesoro + spread)", FIL2)
                t_lbls = [f"t={t:.1f}" for t in tenors_x]
                for j, lbl in enumerate(t_lbls, 2):
                    hdr(ws3, _t_start_row, j, lbl)
                    ws3.column_dimensions[get_column_letter(j)].width = 9
                for i, (rn, row_vals) in enumerate(
                        zip(RATINGS[:_n_spr_rows], allin_x[:_n_spr_rows]), _t_start_row + 1):
                    ws3.cell(row=i, column=1, value=rn).font = Font(bold=True)
                    for j, v in enumerate(row_vals, 2):
                        ws3.cell(row=i, column=j, value=float(v)).number_format = "0.0000%"

                # Sheet 4: Individual bond values
                ws4 = wb.create_sheet("4 Valores Bonos")
                bd_x = _get_bond_data()
                tm_x2 = st.session_state.get("cm_tm", DEFAULT_TM)
                col_off = 1
                for b in bd_x:
                    hdr(ws4, 1, col_off, b["nombre"], FIL2)
                    hdr(ws4, 1, col_off+1, "Prob")
                    hdr(ws4, 1, col_off+2, "Valor ($)")
                    for k in range(3):
                        ws4.column_dimensions[get_column_letter(col_off+k)].width = 14
                    pb = tm_x2[b["rating_idx"]]
                    _nv = len(b["values"])
                    _rl = list(_RATINGS_WORK[:_nv])
                    for ri, (rn, p, v) in enumerate(zip(_rl, pb[:_nv], b["values"]), 2):
                        ws4.cell(row=ri, column=col_off, value=rn)
                        ws4.cell(row=ri, column=col_off+1, value=float(p)).number_format="0.0000%"
                        ws4.cell(row=ri, column=col_off+2, value=float(v)).number_format="#,##0.0000"
                    col_off += 4

                # Sheet 5: Joint distribution
                ws5 = wb.create_sheet("5 Distribucion Conjunta")
                if "cm_df_dist" in st.session_state:
                    df5 = st.session_state["cm_df_dist"].reset_index(drop=True)
                    for j, h in enumerate(df5.columns, 1):
                        hdr(ws5, 1, j, h); ws5.column_dimensions[get_column_letter(j)].width = 24
                    for ri, row in df5.iterrows():
                        for j, v in enumerate(row, 1):
                            c = ws5.cell(row=ri+2, column=j, value=float(v))
                            c.number_format = "#,##0.0000" if j in (1,4) else "0.000000%"
                            c.fill = FIL3
                else:
                    ws5.cell(row=1,column=1, value="Calcula el caso independiente primero.")

                # Sheet 6: VaR / CVaR anual
                ws6 = wb.create_sheet("6 VaR CVaR Anual")
                hdrs6 = ["Conf","Método","VaR 1y ($)","CVaR 1y ($)","E[V] ($)","σ ($)"]
                for j, h in enumerate(hdrs6, 1):
                    hdr(ws6, 1, j, h); ws6.column_dimensions[get_column_letter(j)].width = 20
                rw = 2
                for cf, lb in zip(CONF_LEVELS, CONF_LABELS):
                    if "cm_ivars_param" in st.session_state:
                        r = st.session_state["cm_ivars_param"][cf]
                        for j, v in enumerate([lb,"Independiente",r["VaR"],r["CVaR"],r["EV"],r["sigma"]], 1):
                            c = ws6.cell(row=rw, column=j, value=v)
                            c.fill = FIL3
                            if j >= 3: c.number_format = "#,##0.0000"
                        rw += 1
                    if "cm_cvars" in st.session_state:
                        r = st.session_state["cm_cvars"][cf]
                        for j, v in enumerate([lb,"Correlacionado",r["VaR"],r["CVaR"],r["EV"],r["sigma"]], 1):
                            c = ws6.cell(row=rw, column=j, value=v)
                            if j >= 3: c.number_format = "#,##0.0000"
                        rw += 1

                # Sheet 7: Métricas escaladas (1d · 10d · Capital)
                ws7 = wb.create_sheet("7 Metricas Escaladas")
                ws7.cell(row=1, column=1,
                         value=f"Escala: VaR_1d = VaR_1y ÷ √{TRADING_DAYS}  |  "
                               f"VaR_10d = VaR_1d × √10  |  Capital = 3 × VaR_10d"
                         ).font = Font(italic=True, color="555555")
                hdrs7 = ["Conf","Método",
                         "VaR 1d ($)","CVaR 1d ($)",
                         "VaR 10d ($)","CVaR 10d ($)",
                         "Capital ($)","E[V] ($)"]
                for j, h in enumerate(hdrs7, 1):
                    hdr(ws7, 2, j, h)
                    ws7.column_dimensions[get_column_letter(j)].width = 18
                rw7 = 3
                for metodo, sc_data, fill in [
                    ("Independiente",  st.session_state.get("cm_iscaled_param"), FIL3),
                    ("Correlacionado", st.session_state.get("cm_cscaled"), None),
                ]:
                    if sc_data is None:
                        continue
                    for cf, lb in zip(CONF_LEVELS, CONF_LABELS):
                        r = sc_data[cf]
                        vals_row = [
                            lb, metodo,
                            r["VaR_1d"],  r["CVaR_1d"],
                            r["VaR_10d"], r["CVaR_10d"],
                            r["Capital"], r["EV"],
                        ]
                        for j, v in enumerate(vals_row, 1):
                            cell = ws7.cell(row=rw7, column=j, value=v)
                            if fill: cell.fill = fill
                            if j >= 3: cell.number_format = "#,##0.0000"
                        rw7 += 1

                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                st.session_state["cm_excel"] = buf.getvalue()
                themed_success("Exportación a Excel creada con éxito.")
            except Exception as e:
                themed_error(f"Error procesando la descarga: {e}")

        if "cm_excel" in st.session_state:
            st.download_button(
                "⬇️ Descargar CreditMetrics (.xlsx)",
                data=st.session_state["cm_excel"],
                file_name="CreditMetrics.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )